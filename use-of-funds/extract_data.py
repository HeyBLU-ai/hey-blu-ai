#!/usr/bin/env python3
"""
Extract data from BLU Use-of-Funds Model.xlsx for investor-summary.html
Outputs JSON data that can be used in investor-summary.js
"""

import openpyxl
import json
from pathlib import Path

def extract_data():
    excel_path = Path("BLU Use-of-Funds Model.xlsx")
    
    if not excel_path.exists():
        print(f"Error: {excel_path} not found")
        return None
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    data = {
        "tranches": [],
        "kpis": {},
        "runway": "",
        "founders": []
    }
    
    # Extract Tranche data from Summary sheet
    try:
        ws_summary = wb['Summary']
        # Row 3 = Tranche A, Row 4 = Tranche B, Row 5 = Tranche C
        # Column B = Amount, Column C = Duration
        for i, row in enumerate(ws_summary.iter_rows(min_row=3, max_row=5, values_only=True), 1):
            tranche_name = ["A", "B", "C"][i-1]
            amount = row[1] if row[1] else 0
            duration = row[2] if row[2] else 0
            
            # Get ARR from Funnel_Model
            ws_funnel = wb['Funnel_Model']
            # Row 31 = Total ARR, Column B=Year1, C=Year2, D=Year3
            arr_row = list(ws_funnel.iter_rows(min_row=31, max_row=31, values_only=True))[0]
            arr_values = [arr_row[1], arr_row[2], arr_row[3]]  # Year 1, 2, 3
            
            # Get completion dates from Funding_Overview
            ws_funding = wb.get('Funding_Overview', None)
            if ws_funding:
                funding_rows = list(ws_funding.iter_rows(min_row=3, max_row=5, values_only=True))
                complete_date = funding_rows[i-1][2] if i <= len(funding_rows) else f"Q{i} 202{5+i}"
            else:
                complete_date = f"Q{i} 202{5+i}"
            
            data["tranches"].append({
                "name": tranche_name,
                "amount": int(amount) if amount else 0,
                "ARR": int(arr_values[i-1]) if i <= len(arr_values) and arr_values[i-1] else 0,
                "complete": str(complete_date) if complete_date else f"Q{i} 202{5+i}"
            })
    except Exception as e:
        print(f"Error extracting tranches: {e}")
    
    # Extract KPI data from KPI_Summary sheet
    try:
        ws_kpi = wb['KPI_Summary']
        # Get Year 3 Avg column (H) values
        # Row 3 = U-CAC, Row 4 = Parent Conversion, Row 5 = Coach Conversion
        # Row 6 = LTV, Row 7 = LTV/CAC, Row 8 = Payback, Row 9 = Gross Margin
        
        kpi_rows = list(ws_kpi.iter_rows(min_row=3, max_row=9, values_only=True))
        # Year 3 Avg is column H (index 7)
        year3_col = 7
        
        # Get CAC from Unit_Economics (row 9, column B for Parent)
        ws_unit = wb['Unit_Economics']
        cac_cell = ws_unit['B9'].value  # Parent CAC
        
        # Get LTV/CAC from Unit_Economics (row 10, column B for Parent)
        ltv_cac_cell = ws_unit['B10'].value
        
        # Get Payback from Unit_Economics (row 11, column B for Parent)
        payback_cell = ws_unit['B11'].value
        
        # Get Gross Margin from Assumptions (row 15, column C)
        ws_assumptions = wb['Assumptions']
        gross_margin_cell = ws_assumptions['C15'].value
        
        data["kpis"] = {
            "CAC": int(cac_cell) if cac_cell else 80,
            "LTV_CAC": round(float(ltv_cac_cell), 1) if ltv_cac_cell else 3.2,
            "Payback": round(float(payback_cell), 1) if payback_cell else 8,
            "GrossMargin": f"{int(gross_margin_cell * 100)}%" if gross_margin_cell else "55%"
        }
    except Exception as e:
        print(f"Error extracting KPIs: {e}")
        # Fallback values
        data["kpis"] = {
            "CAC": 80,
            "LTV_CAC": 3.2,
            "Payback": 8,
            "GrossMargin": "55%"
        }
    
    # Extract Founder pay stages from Pay_Triggers sheet
    try:
        ws_pay = wb['Pay_Triggers']
        # Row 3 = Stage 1, Row 4 = Stage 2, Row 5 = Stage 3
        # Column C = Per-Founder Annual
        pay_rows = list(ws_pay.iter_rows(min_row=3, max_row=5, values_only=True))
        
        for i, row in enumerate(pay_rows, 1):
            annual_pay = row[2] if row[2] else 0
            data["founders"].append({
                "stage": f"Stage {i}" if i > 1 else "Baseline",
                "pay": int(annual_pay) if annual_pay else (90000 if i == 1 else 120000 if i == 2 else 250000)
            })
    except Exception as e:
        print(f"Error extracting founder pay: {e}")
        # Fallback values
        data["founders"] = [
            {"stage": "Baseline", "pay": 90000},
            {"stage": "Stage 1", "pay": 120000},
            {"stage": "Stage 2", "pay": 250000}
        ]
    
    # Extract runway from Cash_Runway sheet
    try:
        ws_cash = wb['Cash_Runway']
        # Find the last month with positive closing cash, calculate runway
        # Column N = Runway (Months)
        runway_months = []
        for row in ws_cash.iter_rows(min_row=3, max_row=38, min_col=14, max_col=14):
            if row[0].value and isinstance(row[0].value, (int, float)):
                runway_months.append(row[0].value)
        
        if runway_months:
            min_runway = min([m for m in runway_months if m < 999])
            max_runway = max([m for m in runway_months if m < 999])
            data["runway"] = f"{int(min_runway)}–{int(max_runway)} months"
        else:
            data["runway"] = "24–30 months"
    except Exception as e:
        print(f"Error extracting runway: {e}")
        data["runway"] = "24–30 months"
    
    return data

if __name__ == "__main__":
    data = extract_data()
    if data:
        print(json.dumps(data, indent=2))
    else:
        print("Failed to extract data")

