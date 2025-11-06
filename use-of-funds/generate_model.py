import openpyxl

from openpyxl import Workbook

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from openpyxl.utils import get_column_letter

from openpyxl.formatting.rule import ColorScaleRule

from datetime import datetime



# --- Helper Functions for Styling ---

def setup_workbook():

    """Creates the workbook and all 13 required sheets."""

    wb = Workbook()

    

    # Rename default sheet

    ws_readme = wb.active

    ws_readme.title = "README"

    

    # Create all other sheets in order

    sheets = [

        "Assumptions", "Summary", "Tranche_Detail", "Funnel_Model",

        "Unit_Economics", "Cash_Runway", "Pay_Triggers", "Milestones",

        "Scenarios", "Scenarios_Output", "Sales_Comp", "Infra_Estimates",

        "Financial_Summary", "KPI_Summary", "Funding_Overview"

    ]

    for sheet_name in sheets:

        wb.create_sheet(title=sheet_name)

        

    return wb



def apply_styles(wb):

    """Defines and returns standard styles."""

    styles = {

        'title': Font(bold=True, size=16),

        'header': Font(bold=True, color="FFFFFF"),

        'header_fill': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),

        'subheader': Font(bold=True, size=12),

        'input_fill': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),

        'input_font': Font(color="0000FF", bold=True),

        'currency': openpyxl.styles.NamedStyle(name='currency', number_format='$#,##0'),

        'percent': openpyxl.styles.NamedStyle(name='percent', number_format='0.0%'),

        'integer': openpyxl.styles.NamedStyle(name='integer', number_format='#,##0'),

        'float': openpyxl.styles.NamedStyle(name='float', number_format='0.00'),

    }

    

    # Add named styles to workbook

    if 'currency' not in wb.style_names:

        wb.add_named_style(styles['currency'])

    if 'percent' not in wb.style_names:

        wb.add_named_style(styles['percent'])

    if 'integer' not in wb.style_names:

        wb.add_named_style(styles['integer'])

    if 'float' not in wb.style_names:

        wb.add_named_style(styles['float'])

        

    return styles



def style_sheet(ws, styles, col_widths={}):

    """Applies title, header, and column widths to a sheet."""

    ws['A1'].font = styles['title']

    

    # Apply header style to row 2

    for cell in ws[2]:

        cell.font = styles['header']

        cell.fill = styles['header_fill']

        cell.alignment = Alignment(horizontal="center", vertical="center")

        

    # Apply column widths

    for col, width in col_widths.items():

        ws.column_dimensions[col].width = width



    # Freeze top row

    ws.freeze_panes = 'A3'



def mark_input_cells(ws, cells, styles):

    """Applies blue font and light blue fill to input cells."""

    for cell_ref in cells:

        ws[cell_ref].font = styles['input_font']

        ws[cell_ref].fill = styles['input_fill']



# --- Sheet Population Functions ---



def populate_readme(ws, styles):

    ws['A1'] = "BLU Use-of-Funds Model - README"

    

    ws['A3'] = "How to Use This Model"

    ws['A3'].font = styles['subheader']

    

    instructions = [

        ("1.", "Start on the 'Assumptions' sheet. All key inputs are here."),

        ("2.", "Cells formatted in Blue with a Light Blue Background are inputs. Change these to test different scenarios."),

        ("3.", "All other cells contain formulas and will update automatically."),

        ("4.", "Look at the 'Funnel_Model' tab. This is the engine of the entire spreadsheet. It projects a 3-year forecast starting with 'New Umpires Acquired' as the main input, calculates 'Total Freemium Users Reached' (our viral loop), and then 'Total ARR' based on our B2C conversion rates. All other tabs pull their data from this Funnel_Model."),

        ("5.", "'Tranche_Detail' and 'Cash_Runway' allocate funds and project cash flow."),

        ("6.", "'Unit_Economics' and 'Milestones' track your key performance indicators (KPIs)."),

        ("7.", "'Scenarios' and 'Scenarios_Output' provide a high-level summary of Conservative, Base, and Aggressive outcomes."),

    ]

    

    row = 5

    for item in instructions:

        ws[f'A{row}'] = item[0]

        ws[f'B{row}'] = item[1]

        row += 1

        

    style_sheet(ws, styles, {'A': 5, 'B': 100})



def populate_assumptions(ws, styles):

    ws['A1'] = "Assumptions"

    

    headers = ["Category", "Item", "Base", "Conservative", "Aggressive", "Notes"]

    ws.append(headers)

    

    data = [

        # Market

        ("Market", "Counties Targeted (Year 1)", 5, 3, 8, "Initial focus area"),

        ("Market", "Leagues per County (Avg)", 20, 15, 25, ""),

        ("Market", "Umpires per League (Avg)", 15, 12, 18, ""),

        ("Market", "Parents per Umpire (Avg)", 10, 8, 12, "Parents of players in games umpired"),

        ("Market", "Coaches per League (Avg)", 8, 6, 10, ""),

        # Conversion

        ("Conversion", "Parent App Conversion %", 0.02, 0.01, 0.03, "Freemium to Paid ($199/yr)"),

        ("Conversion", "Coach App Conversion %", 0.03, 0.02, 0.05, "Freemium to Paid ($199/yr)"),

        ("Conversion", "Paid User Churn Rate (Annual)", 0.33, 0.40, 0.25, "Annual churn rate for paid users (inverse of 3yr lifetime)"),

        ("Conversion", "League/Org Conversion % (Yr 1)", 0.00, 0.00, 0.00, "B2C-only strategy - no organization revenue"),

        # Pricing

        ("Pricing", "Coach/Parent App (Annual)", 199, 199, 249, "Annual subscription"),

        ("Pricing", "Organization Price (Annual Avg)", 0, 0, 0, "B2C-only strategy - no organization revenue"),

        # Unit Economics

        ("Unit Economics", "User Lifetime (Years)", 3, 2.5, 3.5, ""),

        ("Unit Economics", "Gross Margin", 0.55, 0.50, 0.60, "Revenue minus COGS (Infra, Support)"),

        ("Unit Economics", "Annual Discount Rate", 0.10, 0.12, 0.08, "For LTV calculation (optional)"),

        # CAC

        ("CAC", "CAC - Ambassador ($/user)", 30, 40, 25, "Paid user (parent/coach)"),

        ("CAC", "CAC - Social/Digital ($/user)", 80, 100, 70, "Paid user (parent/coach)"),

        ("CAC", "CAC - Organization ($/league)", 0, 0, 0, "B2C-only strategy - no organization CAC"),

        # Team & Ops

        ("Team", "Employer Burden", 0.10, 0.12, 0.10, "% on top of base salary (taxes, benefits)"),

        ("Team", "Engineer Salary (Avg)", 120000, 110000, 130000, ""),

        ("Team", "Sales Salary (Avg)", 80000, 75000, 85000, ""),

        # Pay Triggers

        ("Pay Triggers", "Founder Pay - Stage 1 (Annual)", 90000, 90000, 90000, "Annual Salary per Founder ($7.5k/mo)"),

        ("Pay Triggers", "Founder Pay - Stage 2 (Annual)", 120000, 120000, 120000, "Annual Salary per Founder ($10k/mo)"),

        ("Pay Triggers", "Founder Pay - Stage 3 (Annual)", 250000, 250000, 250000, "Annual Salary per Founder"),

        ("Pay Triggers", "MRR Trigger - Stage 2", 10000, 10000, 10000, "MRR threshold for Stage 2 pay"),

        ("Pay Triggers", "MRR Trigger - Stage 3", 50000, 50000, 50000, "MRR threshold for Stage 3 pay"),

        # Tranches

        ("Funding", "Tranche A", 250000, 250000, 250000, "MVP & Pilot"),

        ("Funding", "Tranche B", 750000, 750000, 750000, "Product-Market Fit"),

        ("Funding", "Tranche C", 1000000, 1000000, 1000000, "Growth"),

    ]

    

    input_cells = []

    for i, row_data in enumerate(data, 3):

        ws.append(row_data)

        # Mark Base, Conservative, Aggressive as inputs

        input_cells.extend([f'C{i}', f'D{i}', f'E{i}'])

        

    style_sheet(ws, styles, {'A': 20, 'B': 30, 'C': 15, 'D': 15, 'E': 15, 'F': 40})

    mark_input_cells(ws, input_cells, styles)

    

    # Apply formatting

    for row in ws['C3:E32']:

        for cell in row:

            if ws[f'B{cell.row}'].value and "Conversion" in ws[f'B{cell.row}'].value:

                cell.style = 'percent'

            elif ws[f'B{cell.row}'].value and "Margin" in ws[f'B{cell.row}'].value:

                cell.style = 'percent'

            elif ws[f'B{cell.row}'].value and "Burden" in ws[f'B{cell.row}'].value:

                cell.style = 'percent'

            elif ws[f'B{cell.row}'].value and "Rate" in ws[f'B{cell.row}'].value:

                cell.style = 'percent'

            elif ws[f'A{cell.row}'].value in ["Pricing", "CAC", "Team", "Pay Triggers", "Funding"]:

                 cell.style = 'currency'

            else:

                cell.style = 'integer'



def populate_summary(ws, styles):

    ws['A1'] = "Funding Tranche Summary"

    

    ws.append(["Tranche", "Amount", "Duration (Months)", "Primary Use", "Key Release Milestones"])

    

    data = [

        ("Tranche A", f"=Assumptions!C30", 9, "Build Umpire MVP & Launch Pilots", "Initial Close"),

        ("Tranche B", f"=Assumptions!C31", 12, "Achieve Product-Market Fit (Coach/Parent App)", "Tranche A Milestones Met (5 Leagues, 100 Umpires, $3-5k MRR)"),

        ("Tranche C", f"=Assumptions!C32", 12, "Scale the B2C Engine", "Tranche B Milestones Met ($35k MRR, 2% Conversion)"),

    ]

    

    for row_data in data:

        ws.append(row_data)

        

    ws.append(["Total", "=SUM(B3:B5)", "=SUM(C3:C5)", "", ""])

    

    style_sheet(ws, styles, {'A': 15, 'B': 20, 'C': 20, 'D': 40, 'E': 50})

    

    # Style

    for cell in ws['B']:

        cell.style = 'currency'

    for cell in ws['C']:

        cell.style = 'integer'

    ws['B6'].font = ws['C6'].font = styles['subheader']



def populate_funnel_model(ws, styles):
    """
    Bottom-up funnel model starting with umpires as the sensor network.
    This model builds revenue projections from the ground up, starting with
    the core asset: active umpires capturing data.
    """
    ws['A1'] = "Funnel Model - Bottom-Up Revenue Projections"
    
    # Header row
    ws.append(["Metric", "Year 1", "Year 2", "Year 3", "Total"])
    
    # INPUTS Section
    ws.append(["", "", "", "", ""])  # Spacer
    ws.append(["INPUTS", "", "", "", ""])  # Header row - only label in column A
    ws['A4'].font = styles['subheader']
    
    ws.append(["New Umpires Acquired", 1000, 5000, 15000, "Target from Umpire_Funnel tab"])
    ws.append(["Umpire Churn Rate", 0.20, 0.20, 0.20, "Annual churn rate"])
    ws.append(["Avg. Games per Umpire per Year", 20, 20, 20, "Games captured annually"])
    ws.append(["Avg. Unique Players per Game", 25, 25, 25, "Players reached per game"])
    ws.append(["Avg. Parents/Coaches per Player", 1.5, 1.5, 1.5, "Parents & coaches per player"])
    
    # CALCULATIONS - Bottom-up funnel
    ws.append(["", "", "", "", ""])  # Spacer
    ws.append(["CALCULATIONS", "", "", "", ""])  # Header row - only label in column A
    ws['A13'].font = styles['subheader']
    
    # Active Umpires calculation (with churn)
    # Row mapping: 8=New Umpires (input), 9=Churn, 10=Games, 11=Players, 12=Parents/Coaches
    # Row mapping: 14=New Umpires (calc), 15=Retained, 16=Total Active, 17=Games, 18=Players, 19=Freemium
    ws.append(["New Umpires Acquired", "=B8", "=C8", "=D8", "=D14"])  # Total = Year 3 value (for display)
    ws.append(["Umpires Retained from Prior Year", 0, "=B14*(1-B9)", "=C16*(1-C9)", "=D15"])  # Fixed: use Year 2 total active, not sum of new
    ws.append(["Total Active Umpires", "=B14+B15", "=C14+C15", "=D14+D15", "=D16"])  # Total = Year 3 value (cumulative)
    
    # Games and Reach Calculations
    ws.append(["Total Games Scored", "=B16*B10", "=C16*C10", "=D16*D10", "=D17"])  # Total = Year 3 value
    ws.append(["Total Players Reached", "=B17*B11", "=C17*C11", "=D17*D11", "=D18"])  # Total = Year 3 value
    ws.append(["Total Freemium Users Reached (Parents/Coaches)", "=B18*B12", "=C18*C12", "=D18*D12", "=D19"])  # Total = Year 3 value
    
    # REVENUE CALCULATIONS - Fixed to track cumulative active paid users
    ws.append(["", "", "", "", ""])  # Spacer
    ws.append(["Revenue", "", "", "", ""])  # Header row - only label in column A
    ws['A21'].font = styles['subheader']
    
    # New Conversions (from freemium users reached)
    ws.append(["New Paid Parents", "=B19*Assumptions!C8", "=C19*Assumptions!C8", "=D19*Assumptions!C8", "=D22"])  # Total = Year 3 value
    ws.append(["New Paid Coaches", "=B19*Assumptions!C9", "=C19*Assumptions!C9", "=D19*Assumptions!C9", "=D23"])  # Total = Year 3 value
    
    # Paid User Retention (accounting for churn)
    # Row mapping: 22=New Paid Parents, 23=New Paid Coaches, 24=Retained Parents, 25=Retained Coaches, 26=Total Parents, 27=Total Coaches
    # Year 2: Retain from Year 1 new paid users
    # Year 3: Retain from Year 2 cumulative active paid users (not just new)
    ws.append(["Paid Parents Retained from Prior Year", 0, "=B22*(1-Assumptions!C10)", "=C26*(1-Assumptions!C10)", "=D24"])  # C10 = Paid User Churn Rate
    ws.append(["Paid Coaches Retained from Prior Year", 0, "=B23*(1-Assumptions!C10)", "=C27*(1-Assumptions!C10)", "=D25"])  # C10 = Paid User Churn Rate
    
    # Cumulative Active Paid Users
    ws.append(["Total Active Paid Parents", "=B22+B24", "=C22+C24", "=D22+D24", "=D26"])  # Total = Year 3 value
    ws.append(["Total Active Paid Coaches", "=B23+B25", "=C23+C25", "=D23+D25", "=D27"])  # Total = Year 3 value
    
    # Revenue (from cumulative active paid users, not just new)
    ws.append(["Parent App Revenue", "=B26*Assumptions!C11", "=C26*Assumptions!C11", "=D26*Assumptions!C11", "=D28"])  # Total = Year 3 value
    ws.append(["Coach App Revenue", "=B27*Assumptions!C11", "=C27*Assumptions!C11", "=D27*Assumptions!C11", "=D29"])  # Total = Year 3 value
    ws.append(["Total ARR", "=B29+B30", "=C29+C30", "=D29+D30", "=D31"])  # Total = Year 3 value
    ws.append(["Total MRR (Avg Year 3)", "", "", "=D31/12", ""])
    
    style_sheet(ws, styles, {'A': 35, 'B': 18, 'C': 18, 'D': 18, 'E': 25})
    
    # Formatting
    # Format INPUTS as blue (editable)
    for row in range(8, 13):
        for col in ['B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.fill = styles['input_fill']
            cell.font = styles['input_font']
            if row == 9:  # Churn rate row
                cell.style = 'percent'
            else:
                cell.style = 'integer' if row in [8, 10, 11] else 'float'
    
    # Format calculated values
    for row in [14, 15, 16, 17, 18, 19, 22, 23, 24, 25, 26, 27]:
        for col in ['B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            cell.style = 'integer'
    
    # Format revenue values as currency
    for row in [28, 29, 30, 31]:
        for col in ['B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            if cell.value and str(cell.value).startswith('='):
                cell.style = 'currency'
    
    # Add note about top-of-funnel
    ws['A19'].font = styles['subheader']
    # Note: Comment removed - Excel doesn't require it for functionality





def populate_tranche_detail(ws, styles):

    ws['A1'] = "Tranche Detail - Use of Funds"

    

    ws.append(["Category", "Item", "Tranche A", "Tranche B", "Tranche C", "Total"])



    data = [

        # Payroll
        # Note: Row numbers in Assumptions sheet = data index + 3 (row 1=title, row 2=headers, row 3+=data)
        # Employer Burden = index 16 → row 19 → Column C = C19
        # Engineer Salary = index 17 → row 20 → Column C = C20
        # Sales Salary = index 18 → row 21 → Column C = C21
        # Founder Stage 1 = index 19 → row 22 → Column C = C22 ($90k annual = $7.5k/month - CURRENT PAY)
        # Founder Stage 2 = index 20 → row 23 → Column C = C23 ($120k annual = $10k/month - triggers at $10k MRR)
        # Summary sheet: Row 3=Tranche A (9 months), Row 4=Tranche B (12 months), Row 5=Tranche C (12 months)
        # Founder Pay Progression (scales UP with milestones, stays at Stage 2 through Tranche C):
        # Tranche A: Stage 1 ($90k annual × 2 founders × 9/12 months) = $135k
        # Tranche B: Stage 1 for first 6 months, Stage 2 for last 6 months (when $10k MRR milestone hit)
        # Tranche C: Stage 2 for full period ($120k annual × 2 founders × 12/12 months) = $240k
        # This demonstrates founder alignment - prioritizing company growth and capital efficiency over personal compensation

        ("Payroll", "Founders (2)", f"=Assumptions!C22*2*(Summary!C3/12)", f"=(Assumptions!C22*2*(6/12))+(Assumptions!C23*2*(6/12))", f"=Assumptions!C23*2*(Summary!C5/12)", "=SUM(C3:E3)"),

        ("Payroll", "Engineers (2)", f"=Assumptions!C20*2*(1+Assumptions!C19)*(Summary!C3/12)", f"=Assumptions!C20*2*(1+Assumptions!C19)*(Summary!C4/12)", f"=Assumptions!C20*2*(1+Assumptions!C19)*(Summary!C5/12)", "=SUM(C4:E4)"),

        ("Payroll", "Growth & Partnerships (1 -> 2)", 0, f"=Assumptions!C21*1*(1+Assumptions!C19)*(Summary!C4/12)", f"=Assumptions!C21*1*(1+Assumptions!C19)*((Summary!C5-3)/12)+Assumptions!C21*2*(1+Assumptions!C19)*(3/12)", "=SUM(C5:E5)"),

        ("Payroll", "Total Payroll", "=SUM(C3:C5)", "=SUM(D3:D5)", "=SUM(E3:E5)", "=SUM(F3:F5)"),

        # S&M

        # S&M - Dynamic CAC based on Funnel_Model new paid users
        # Formula: (New Paid Parents + New Paid Coaches) × CAC per user from Assumptions
        # Tranche A uses Year 1, Tranche B uses Year 2, Tranche C uses Year 3
        ("Sales & Marketing", "CAC - Ambassador", f"=(Funnel_Model!B22+Funnel_Model!B23)*Assumptions!C17", f"=(Funnel_Model!C22+Funnel_Model!C23)*Assumptions!C17", f"=(Funnel_Model!D22+Funnel_Model!D23)*Assumptions!C17", "=SUM(C7:E7)"),

        ("Sales & Marketing", "CAC - Social/Digital", f"=(Funnel_Model!B22+Funnel_Model!B23)*Assumptions!C18", f"=(Funnel_Model!C22+Funnel_Model!C23)*Assumptions!C18", f"=(Funnel_Model!D22+Funnel_Model!D23)*Assumptions!C18", "=SUM(C8:E8)"),

        ("Sales & Marketing", "CAC - Organization", 0, 0, 0, "=SUM(C9:E9)"),

        ("Sales & Marketing", "Total S&M", "=SUM(C7:C9)", "=SUM(D7:D9)", "=SUM(E7:E9)", "=SUM(F7:F9)"),

        # R&D

        ("R&D", "Software & Services", 15000, 30000, 40000, "=SUM(C11:E11)"),

        ("R&D", "Total R&D", "=C11", "=D11", "=E11", "=F11"),

        # Infra

        ("Infra", "Hosting & Data", f"=Infra_Estimates!F6*(Summary!C3/12)", f"=Infra_Estimates!F6", f"=Infra_Estimates!F6*1.5", "=SUM(C13:E13)"), # Simplified

        ("Infra", "Total Infra", "=C13", "=D13", "=E13", "=F13"),

        # G&A

        ("G&A", "Legal & Admin", 20000, 30000, 40000, "=SUM(C15:E15)"),

        ("G&A", "Office & Misc", 10000, 20000, 20000, "=SUM(C16:E16)"),

        ("G&A", "Total G&A", "=SUM(C15:C16)", "=SUM(D15:D16)", "=SUM(E15:E16)", "=SUM(F15:F16)"),

        # Totals
        # Row structure: 1=title, 2=headers, 3=Founders, 4=Engineers, 5=Sales, 6=Payroll Total,
        # 7=Ambassador, 8=Social, 9=Org, 10=S&M Total, 11=Software, 12=R&D Total,
        # 13=Hosting, 14=Infra Total, 15=Legal, 16=Office, 17=G&A Total,
        # 18=(blank), 19=Total Allocation, 20=Tranche Amount, 21=Surplus

        ("", "", "", "", "", ""),

        ("Total Allocation", "", "=SUM(C6,C10,C12,C14,C17)", "=SUM(D6,D10,D12,D14,D17)", "=SUM(E6,E10,E12,E14,E17)", "=SUM(F6,F10,F12,F14,F17)"),

        ("Tranche Amount", "", f"=Assumptions!C30", f"=Assumptions!C31", f"=Assumptions!C32", "=SUM(C20:E20)"),

        ("Surplus / Deficit", "", "=C20-C19", "=D20-D19", "=E20-E19", "=F20-F19"),

    ]

    

    for row_data in data:

        ws.append(row_data)



    style_sheet(ws, styles, {'A': 20, 'B': 25, 'C': 18, 'D': 18, 'E': 18, 'F': 20})



    # Formatting

    for row in ws.iter_rows(min_row=3, max_row=25, min_col=3, max_col=6):

        for cell in row:

            cell.style = 'currency'

            

    for row in [6, 11, 14, 17, 21, 23, 24, 25]:

        ws[f'A{row}'].font = ws[f'B{row}'].font = ws[f'C{row}'].font = ws[f'D{row}'].font = ws[f'E{row}'].font = ws[f'F{row}'].font = styles['subheader']



def populate_unit_economics(ws, styles):

    ws['A1'] = "Unit Economics (Base Scenario)"

    

    ws.append(["Metric", "Parent (App)", "Coach (App)", "Organization (League)", "Blended"])

    

    data = [

        ("ARPU (Annual)", f"=Assumptions!C11", f"=Assumptions!C11", f"=Assumptions!C12", f"=Funnel_Model!D31 / (Funnel_Model!D27 + Funnel_Model!D28)"),

        ("Gross Margin", f"=Assumptions!C15", f"=Assumptions!C15", f"=Assumptions!C15", f"=Assumptions!C15"),

        ("Contribution Margin (Annual)", "=B3*B4", "=C3*C4", "=D3*D4", "=E3*E4"),

        ("Lifetime (Years)", f"=Assumptions!C14", f"=Assumptions!C14", f"=Assumptions!C14", f"=Assumptions!C14"),

        ("LTV", "=B5*B6", "=C5*C6", "=D5*D6", "=E5*E6"),

        ("", "", "", "", ""), # Spacer

        ("CAC", f"=Assumptions!C18", f"=Assumptions!C18", f"=Assumptions!C19", "N/A"), # Simplified CAC

        ("LTV / CAC Ratio", "=B7/B9", "=C7/C9", "=D7/D9", "N/A"),

        ("Payback Period (Months)", f"=(B9/B5)*12", f"=(C9/C5)*12", f"=(D9/D5)*12", "N/A"),

    ]

    

    for row_data in data:

        ws.append(row_data)



    style_sheet(ws, styles, {'A': 30, 'B': 18, 'C': 18, 'D': 20, 'E': 18})



    # Formatting

    for row in [3, 5, 7, 9]:

        for col in ['B', 'C', 'D', 'E']:

            ws[f'{col}{row}'].style = 'currency'

    ws['B4'].style = ws['C4'].style = ws['D4'].style = ws['E4'].style = 'percent'

    for row in [6, 10]:

        for col in ['B', 'C', 'D', 'E']:

            ws[f'{col}{row}'].style = 'float'

    ws['B11'].style = ws['C11'].style = ws['D11'].style = ws['E11'].style = 'float'

    

    # Color scale for LTV/CAC

    ws.conditional_formatting.add('B10:D10', ColorScaleRule(start_type='num', start_value=1, start_color='FF0000', mid_type='num', mid_value=3, mid_color='FFFF00', end_type='num', end_value=5, end_color='00B050'))



def populate_cash_runway(ws, styles):

    ws['A1'] = "Monthly Cash Runway"

    

    headers = [

        "Month", "Opening Cash", "Financing", "Revenue", "Total Cash In",

        "Payroll", "S&M", "R&D", "Infra", "G&A", "Total Cash Out",

        "Net Burn", "Closing Cash", "Runway (Months)"

    ]

    ws.append(headers)



    # Simplified monthly costs from tranches

    ws['O3'] = "Tranche A Monthly Burn"

    ws['P3'] = f"=(Tranche_Detail!C19) / Summary!C3"

    ws['O4'] = "Tranche B Monthly Burn"

    ws['P4'] = f"=(Tranche_Detail!D19) / Summary!C4"

    ws['O5'] = "Tranche C Monthly Burn"

    ws['P5'] = f"=(Tranche_Detail!E19) / Summary!C5"

    

    ws['O7'] = "Tranche A Monthly Payroll"

    ws['P7'] = f"=(Tranche_Detail!C6) / Summary!C3"

    ws['O8'] = "Tranche A Monthly S&M"

    ws['P8'] = f"=(Tranche_Detail!C10) / Summary!C3"

    # ... and so on for all categories

    

    # Month 1

    ws['A3'] = 1

    ws['B3'] = 0

    ws['C3'] = f"=Assumptions!C30" # Tranche A

    ws['D3'] = 0 # No revenue M1

    ws['E3'] = "=C3+D3"

    ws['F3'] = f"=(Tranche_Detail!C6) / Summary!C3"  # Payroll T-A

    ws['G3'] = f"=(Tranche_Detail!C10) / Summary!C3" # S&M T-A

    ws['H3'] = f"=(Tranche_Detail!C12) / Summary!C3" # R&D T-A

    ws['I3'] = f"=(Tranche_Detail!C14) / Summary!C3" # Infra T-A

    ws['J3'] = f"=(Tranche_Detail!C17) / Summary!C3" # G&A T-A

    ws['K3'] = "=SUM(F3:J3)"

    ws['L3'] = "=K3-D3" # Net Burn

    ws['M3'] = "=B3+E3-K3"

    ws['N3'] = f"=IF(L3>0, M3/L3, 999)"



    # Loop for 36 months

    for m in range(2, 37):

        row = m + 2

        ws[f'A{row}'] = m

        ws[f'B{row}'] = f"=M{row-1}" # Opening Cash

        

        # Financing Triggers

        if m == 10: # Tranche B trigger (Month 10)

             ws[f'C{row}'] = f"=Assumptions!C31"

        elif m == 22: # Tranche C trigger (Month 22)

            ws[f'C{row}'] = f"=Assumptions!C32"

        else:

            ws[f'C{row}'] = 0

            

        # Revenue (simple linear ramp to Year 1 ARR)

        ws[f'D{row}'] = f"=IF(A{row}<=12, (Funnel_Model!B31/12)*(A{row}/12), IF(A{row}<=24, (Funnel_Model!C31/12), (Funnel_Model!D31/12)))"

        

        ws[f'E{row}'] = f"=C{row}+D{row}"

        

        # Determine which tranche's burn rate to use

        payroll_formula = f"=IF(A{row}<=Summary!C3, (Tranche_Detail!C6)/Summary!C3, IF(A{row}<=(Summary!C3+Summary!C4), (Tranche_Detail!D6)/Summary!C4, (Tranche_Detail!E6)/Summary!C5))"

        sm_formula = f"=IF(A{row}<=Summary!C3, (Tranche_Detail!C10)/Summary!C3, IF(A{row}<=(Summary!C3+Summary!C4), (Tranche_Detail!D10)/Summary!C4, (Tranche_Detail!E10)/Summary!C5))"

        rd_formula = f"=IF(A{row}<=Summary!C3, (Tranche_Detail!C12)/Summary!C3, IF(A{row}<=(Summary!C3+Summary!C4), (Tranche_Detail!D12)/Summary!C4, (Tranche_Detail!E12)/Summary!C5))"

        infra_formula = f"=IF(A{row}<=Summary!C3, (Tranche_Detail!C14)/Summary!C3, IF(A{row}<=(Summary!C3+Summary!C4), (Tranche_Detail!D14)/Summary!C4, (Tranche_Detail!E14)/Summary!C5))"

        ga_formula = f"=IF(A{row}<=Summary!C3, (Tranche_Detail!C17)/Summary!C3, IF(A{row}<=(Summary!C3+Summary!C4), (Tranche_Detail!D17)/Summary!C4, (Tranche_Detail!E17)/Summary!C5))"



        ws[f'F{row}'] = payroll_formula

        ws[f'G{row}'] = sm_formula

        ws[f'H{row}'] = rd_formula

        ws[f'I{row}'] = infra_formula

        ws[f'J{row}'] = ga_formula

        

        ws[f'K{row}'] = f"=SUM(F{row}:J{row})"

        ws[f'L{row}'] = f"=K{row}-D{row}"

        ws[f'M{row}'] = f"=B{row}+E{row}-K{row}"

        ws[f'N{row}'] = f"=IF(L{row}>0, M{row}/L{row}, 999)"



    style_sheet(ws, styles, {'A': 8, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 18, 'I': 18, 'J': 18, 'K': 18, 'L': 18, 'M': 18, 'N': 18})

    

    # Formatting

    for row in ws.iter_rows(min_row=3, max_row=38, min_col=2, max_col=13):

        for cell in row:

            cell.style = 'currency'

    for row in ws.iter_rows(min_row=3, max_row=38, min_col=14, max_col=14):

        for cell in row:

            cell.style = 'float'

            

def populate_pay_triggers(ws, styles):

    ws['A1'] = "Founder Compensation (OpEx)"

    ws['A1'].font = styles['title']

    

    ws.append(["Stage", "MRR Trigger", "Per-Founder Annual", "Combined Annual Cost", "Monthly Impact"])

    ws[2][0].font = styles['header']

    ws[2][1].font = styles['header']

    ws[2][2].font = styles['header']

    ws[2][3].font = styles['header']

    ws[2][4].font = styles['header']

    for cell in ws[2]:

        cell.fill = styles['header_fill']

        cell.alignment = Alignment(horizontal="center", vertical="center")

    

    data = [

        (1, 0, f"=Assumptions!C22", f"=Assumptions!C22*2", f"=Assumptions!C22*2/12"),

        (2, f"=Assumptions!C25", f"=Assumptions!C23", f"=Assumptions!C23*2", f"=Assumptions!C23*2/12"),

        (3, f"=Assumptions!C26", f"=Assumptions!C24", f"=Assumptions!C24*2", f"=Assumptions!C24*2/12"),

    ]

    

    for row_data in data:

        ws.append(row_data)

    # Format currency columns

    for row in range(3, 6):

        for col in ['C', 'D', 'E']:

            ws[f'{col}{row}'].style = 'currency'

    

    style_sheet(ws, styles, {'A': 30, 'B': 25, 'C': 18, 'D': 18, 'E': 30})

    

    # Formatting

    for col in ['B', 'C']:

        for row in [3, 4, 5, 8, 9, 10]:

            ws[f'{col}{row}'].style = 'currency'

            

def populate_milestones(ws, styles):

    ws['A1'] = "Milestones & KPIs"

    

    ws.append(["Tranche", "Milestone", "Metric", "Target", "Current", "Status", "Completion Date"])

    

    # Calculate completion dates based on tranche start + duration
    # Assuming Tranche A starts at initial close, Tranche B starts after A completes, etc.
    data = [

        ("A", "MVP Launch", "App Live", 1, 1, "=IF(E3>=D3,\"Met\",\"Pending\")", "Apr 2025"),

        ("A", "Umpires Onboarded", "Users", 100, 0, "=IF(E4>=D4,\"Met\",\"Pending\")", "Jun 2025"),

        ("A", "Pilot Leagues Secured", "Leagues", 5, 0, "=IF(E5>=D5,\"Met\",\"Pending\")", "Jul 2025"),  # Note: League tracking may need separate model

        ("A", "Generate $3-5k MRR", "MRR", 4000, f"=Funnel_Model!D32", "=IF(E6>=D6,\"Met\",\"Pending\")", "Jan 2026"),

        

        ("B", "Tranche A Milestones Met", "Status", 1, f"=IF(AND(F5=\"Met\",F6=\"Met\"),1,0)", "=IF(E7>=D7,\"Met\",\"Pending\")", "Jan 2026"),

        ("B", "Achieve $35k MRR", "MRR", 35000, f"=Funnel_Model!D32", "=IF(E8>=D8,\"Met\",\"Pending\")", "Dec 2026"),

        ("B", "Parent Conversion", "%", f"=Assumptions!C8", f"=Assumptions!C8", "=IF(E9>=D9,\"Met\",\"Pending\")", "Dec 2026"),

        ("B", "Coach Conversion", "%", f"=Assumptions!C9", f"=Assumptions!C9", "=IF(E10>=D10,\"Met\",\"Pending\")", "Dec 2026"),

        

        ("C", "Tranche B Milestones Met", "Status", 1, f"=IF(AND(F8=\"Met\",F9=\"Met\",F10=\"Met\"),1,0)", "=IF(E11>=D11,\"Met\",\"Pending\")", "Dec 2026"),

        ("C", "Achieve $100k MRR", "MRR", 100000, f"=Funnel_Model!D32", "=IF(E12>=D12,\"Met\",\"Pending\")", "Dec 2027"),

        ("C", "Org Contracts", "Count", 10, 0, "=IF(E13>=D13,\"Met\",\"Pending\")", "Dec 2027"),  # Note: Organization revenue tracking may need separate model

    ]

    

    for row_data in data:

        ws.append(row_data)



    style_sheet(ws, styles, {'A': 10, 'B': 30, 'C': 20, 'D': 15, 'E': 15, 'F': 15})

    

    # Formatting

    ws['D8'].style = ws['E8'].style = 'percent'

    ws['D9'].style = ws['E9'].style = 'percent'

    ws['D7'].style = ws['E7'].style = 'currency'

    ws['D11'].style = ws['E11'].style = 'currency'

    mark_input_cells(ws, ['E3', 'E4'], styles) # Mark 'Current' as manual inputs for non-formula items



def populate_scenarios(ws, styles):

    ws['A1'] = "Scenario Assumptions"

    

    ws.append(["Variable", "Conservative", "Base", "Aggressive"])

    

    data = [

        ("Parent Conversion %", f"=Assumptions!D8", f"=Assumptions!C8", f"=Assumptions!E8"),

        ("Coach Conversion %", f"=Assumptions!D9", f"=Assumptions!C9", f"=Assumptions!E9"),

        ("Org Conversion %", f"=Assumptions!D10", f"=Assumptions!C10", f"=Assumptions!E10"),

        ("Org Price (Annual Avg)", f"=Assumptions!D12", f"=Assumptions!C12", f"=Assumptions!E12"),

        ("CAC - Ambassador", f"=Assumptions!D17", f"=Assumptions!C17", f"=Assumptions!E17"),

        ("CAC - Social/Digital", f"=Assumptions!D18", f"=Assumptions!C18", f"=Assumptions!E18"),

        ("CAC - Organization", f"=Assumptions!D19", f"=Assumptions!C19", f"=Assumptions!E19"),

    ]

    

    for row_data in data:

        ws.append(row_data)

        

    style_sheet(ws, styles, {'A': 30, 'B': 18, 'C': 18, 'D': 18})



    # Formatting

    for row in [3, 4, 5]:

        for col in ['B', 'C', 'D']:

            ws[f'{col}{row}'].style = 'percent'

    for row in [6, 7, 8, 9]:

        for col in ['B', 'C', 'D']:

            ws[f'{col}{row}'].style = 'currency'



def populate_scenarios_output(ws, styles):

    ws['A1'] = "Scenarios Output Summary"

    

    ws.append(["Metric", "Conservative", "Base", "Aggressive"])

    

    # These formulas are re-calculations of the model based on scenario inputs

    # This is a simplified version. A full model would use CHOOSE or OFFSET.

    

    # Year 3 ARR

    ws['A3'] = "Year 3 ARR"

    ws['B3'] = f"=(Funnel_Model!D27*Scenarios!B3*Assumptions!D11) + (Funnel_Model!D28*Scenarios!B4*Assumptions!D11)"  # Year 3 active paid users with scenario multipliers

    ws['C3'] = f"=Funnel_Model!D31"  # Base Year 3 ARR

    ws['D3'] = f"=(Funnel_Model!D27*Scenarios!D3*Assumptions!E11) + (Funnel_Model!D28*Scenarios!D4*Assumptions!E11)"  # Year 3 active paid users with scenario multipliers

    

    # Year 3 Total Paid Users

    ws['A4'] = "Year 3 Total Paid Users"

    ws['B4'] = f"=(Funnel_Model!D27*Scenarios!B3) + (Funnel_Model!D28*Scenarios!B4)"  # Year 3 active paid users with scenario multipliers

    ws['C4'] = f"=Funnel_Model!D27 + Funnel_Model!D28"  # Base: Year 3 total active paid users

    ws['D4'] = f"=(Funnel_Model!D27*Scenarios!D3) + (Funnel_Model!D28*Scenarios!D4)"  # Year 3 active paid users with scenario multipliers

    

    # Blended LTV

    ws['A5'] = "Blended LTV (Year 3)"

    ws['B5'] = f"=(B3/B4) * Assumptions!D14 * Assumptions!D13"

    ws['C5'] = f"=Unit_Economics!E7"

    ws['D5'] = f"=(D3/D4) * Assumptions!E14 * Assumptions!E13"

    

    # Blended Payback (Months)

    ws['A6'] = "Blended Payback (Months) - Est."

    ws['B6'] = f"( (Scenarios!B7*B4) ) / (B3*Assumptions!D14) * 12"  # Very simplified - removed org contracts reference

    ws['C6'] = f"=Unit_Economics!E11"

    ws['D6'] = f"( (Scenarios!D7*D4) ) / (D3*Assumptions!E14) * 12"  # Very simplified - removed org contracts reference

    

    style_sheet(ws, styles, {'A': 30, 'B': 18, 'C': 18, 'D': 18})

    

    # Formatting

    for row in [3, 5]:

        for col in ['B', 'C', 'D']:

            ws[f'{col}{row}'].style = 'currency'

    for row in [4]:

        for col in ['B', 'C', 'D']:

            ws[f'{col}{row}'].style = 'integer'

    for row in [6]:

        for col in ['B', 'C', 'D']:

            ws[f'{col}{row}'].style = 'float'



def populate_sales_comp(ws, styles):

    ws['A1'] = "Sales Comp Model (Example)"

    

    ws.append(["Role", "Base Salary", "Commission Rate", "Bonus Target", "Notes / Triggers"])

    

    data = [

        ("Org Sales Rep (Tranche B)", f"=Assumptions!C21", 0.10, 20000, "10% of 1st Year ACV. Bonus on 100% quota."),

        ("Partner Manager (Tranche C)", 90000, 0.05, 15000, "5% of ACV from league partners."),

        ("Head of Sales (Tranche C)", 140000, 0.02, 50000, "2% team override. Bonus on team quota."),

    ]

    

    for row_data in data:

        ws.append(row_data)



    style_sheet(ws, styles, {'A': 30, 'B': 18, 'C': 18, 'D': 18, 'E': 40})

    

    # Formatting

    ws['B3'].style = ws['B4'].style = ws['B5'].style = 'currency'

    ws['D3'].style = ws['D4'].style = ws['D5'].style = 'currency'

    ws['C3'].style = ws['C4'].style = ws['C5'].style = 'percent'



def populate_infra_estimates(ws, styles):

    ws['A1'] = "Infrastructure Cost Estimates"

    

    ws.append(["Item", "Cost Driver", "Cost per Unit", "Units (Year 2)", "Total Annual Cost", "Notes"])

    

    data = [

        ("Umpire Onboarding/Training", "Per Umpire", 10, f"=Funnel_Model!C16", "=C3*D3", "Manuals, support, background check subsidy"),

        ("Audio Data Storage (Cloud)", "Per Game (Avg)", 0.50, f"=Funnel_Model!C17", "=C4*D4", "Total games scored in Year 2"),

        ("Data Processing (AI)", "Per Game (Avg)", 1.00, f"=D4", "=C5*D5", "Analytics & transcription costs"),

    ]

    

    for row_data in data:

        ws.append(row_data)



    ws['A6'] = "Total Annual Infra COGS"

    ws['E6'] = "=SUM(E3:E5)"

    ws['E6'].font = styles['subheader']

    

    ws['A8'] = "Adjusted Gross Margin"

    ws['A8'].font = styles['subheader']

    ws['A9'] = "Base Gross Margin"

    ws['B9'] = f"=Assumptions!C15"

    ws['A10'] = "Infra COGS as % of Revenue (Y2)"

    ws['B10'] = f"=E6/Funnel_Model!C31"

    ws['A11'] = "Adjusted Gross Margin"

    ws['B11'] = "=B9-B10"

    

    style_sheet(ws, styles, {'A': 30, 'B': 20, 'C': 15, 'D': 18, 'E': 20, 'F': 40})

    

    # Formatting

    ws['C3'].style = ws['C4'].style = ws['C5'].style = 'currency'

    ws['E3'].style = ws['E4'].style = ws['E5'].style = ws['E6'].style = 'currency'

    ws['D3'].style = ws['D4'].style = ws['D5'].style = 'integer'

    ws['B9'].style = ws['B10'].style = ws['B11'].style = 'percent'




def populate_financial_summary(ws, styles):
    """Top-line financial summary showing capital-to-ARR ratio by tranche."""
    ws['A1'] = "Financial Summary - Capital Efficiency"
    ws['A1'].font = styles['title']
    
    ws.append(["Tranche", "Capital Invested", "Projected ARR", "Projected MRR", "Capital-to-ARR Ratio"])
    
    # Header formatting
    for cell in ws[2]:
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Tranche A: 9 months, Year 1 ARR
    ws.append(["Tranche A", f"=Assumptions!C30", f"=Funnel_Model!B31", f"=Funnel_Model!B31/12", f"=B3/C3"])
    
    # Tranche B: Cumulative (A+B), Year 2 ARR
    ws.append(["Tranche B", f"=Assumptions!C30+Assumptions!C31", f"=Funnel_Model!C31", f"=Funnel_Model!C31/12", f"=B4/C4"])
    
    # Tranche C: Total ($2M), Year 3 ARR
    ws.append(["Tranche C", f"=Assumptions!C30+Assumptions!C31+Assumptions!C32", f"=Funnel_Model!D31", f"=Funnel_Model!D31/12", f"=B5/C5"])
    
    # Totals row
    ws.append(["", "", "", "", ""])
    ws['A7'] = "Total"
    ws['A7'].font = styles['subheader']
    ws['B7'] = "=B5"
    ws['C7'] = "=C5"
    ws['D7'] = "=D5"
    ws['E7'] = "=B7/C7"
    
    # Formatting
    for row in range(3, 8):
        for col in ['B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            if col == 'E':  # Ratio column
                cell.number_format = '0.00'
            else:
                cell.style = 'currency'
    
    style_sheet(ws, styles, {'A': 20, 'B': 18, 'C': 18, 'D': 18, 'E': 18})


def populate_kpi_summary(ws, styles):
    """Quarterly KPI dashboard showing key metrics."""
    ws['A1'] = "KPI Summary - Quarterly Metrics"
    ws['A1'].font = styles['title']
    
    ws.append(["Metric", "Q1", "Q2", "Q3", "Q4", "Year 1 Avg", "Year 2 Avg", "Year 3 Avg"])
    
    # Header formatting
    for cell in ws[2]:
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # U-CAC (Umpire Acquisition Cost) - simplified to annual
    ws.append(["U-CAC ($/umpire)", "=Assumptions!C17", "=Assumptions!C17", "=Assumptions!C17", "=Assumptions!C17", "=Assumptions!C17", "=Assumptions!C17", "=Assumptions!C17"])
    
    # Conversion % - Parent (row 3)
    ws.append(["Parent Conversion %", f"=Assumptions!C8", f"=Assumptions!C8", f"=Assumptions!C8", f"=Assumptions!C8", f"=Assumptions!C8", f"=Assumptions!C8", f"=Assumptions!C8"])
    
    # Conversion % - Coach (row 4)
    ws.append(["Coach Conversion %", f"=Assumptions!C9", f"=Assumptions!C9", f"=Assumptions!C9", f"=Assumptions!C9", f"=Assumptions!C9", f"=Assumptions!C9", f"=Assumptions!C9"])
    
    # ARPU (row 5)
    ws.append(["ARPU (Annual)", f"=Assumptions!C11", f"=Assumptions!C11", f"=Assumptions!C11", f"=Assumptions!C11", f"=Assumptions!C11", f"=Assumptions!C11", f"=Unit_Economics!E3"])
    
    # LTV (row 6)
    ws.append(["LTV", f"=Unit_Economics!B7", f"=Unit_Economics!B7", f"=Unit_Economics!B7", f"=Unit_Economics!B7", f"=Unit_Economics!B7", f"=Unit_Economics!B7", f"=Unit_Economics!E7"])
    
    # LTV/CAC Ratio (row 7)
    ws.append(["LTV/CAC Ratio", f"=Unit_Economics!B10", f"=Unit_Economics!B10", f"=Unit_Economics!B10", f"=Unit_Economics!B10", f"=Unit_Economics!B10", f"=Unit_Economics!B10", f"=Unit_Economics!E10"])
    
    # Payback Period (Months) (row 8)
    ws.append(["Payback Period (Months)", f"=Unit_Economics!B11", f"=Unit_Economics!B11", f"=Unit_Economics!B11", f"=Unit_Economics!B11", f"=Unit_Economics!B11", f"=Unit_Economics!B11", f"=Unit_Economics!E11"])
    
    # Gross Margin (row 9)
    ws.append(["Gross Margin", f"=Assumptions!C15", f"=Assumptions!C15", f"=Assumptions!C15", f"=Assumptions!C15", f"=Assumptions!C15", f"=Assumptions!C15", f"=Assumptions!C15"])
    
    # Formatting
    for row in range(3, 11):
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            cell = ws[f'{col}{row}']
            if row in [3, 4]:  # Conversion % rows (Parent and Coach)
                cell.style = 'percent'
            elif row == 5:  # ARPU
                cell.style = 'currency'
            elif row == 6:  # LTV
                cell.style = 'currency'
            elif row == 7:  # LTV/CAC
                cell.number_format = '0.00'
            elif row == 8:  # Payback
                cell.number_format = '0.0'
            elif row == 9:  # Gross Margin
                cell.style = 'percent'
            elif row == 2:  # U-CAC
                cell.style = 'currency'
    
    style_sheet(ws, styles, {'A': 25, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15})


def populate_funding_overview(ws, styles):
    """Timeline overview of funding tranches with deliverables."""
    ws['A1'] = "Funding Overview - Timeline"
    ws['A1'].font = styles['title']
    
    ws.append(["Tranche", "Amount", "Start Date", "End Date", "Duration", "Major Deliverable"])
    
    # Header formatting
    for cell in ws[2]:
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Tranche A: MVP
    ws.append(["Tranche A → MVP", f"=Assumptions!C30", "Jan 2025", "Sep 2025", "9 months", "Umpire MVP, 100+ umpires, $3-5k MRR"])
    
    # Tranche B: Pilot Scale
    ws.append(["Tranche B → Pilot Scale", f"=Assumptions!C31", "Jan 2026", "Dec 2026", "12 months", "$35k MRR, 2% parent / 3% coach conversion"])
    
    # Tranche C: Growth
    ws.append(["Tranche C → Growth", f"=Assumptions!C32", "Jan 2027", "Dec 2027", "12 months", "$100k MRR, scalable B2C engine"])
    
    # Total
    ws.append(["", "", "", "", "", ""])
    ws['A7'] = "Total"
    ws['A7'].font = styles['subheader']
    ws['B7'] = f"=Assumptions!C30+Assumptions!C31+Assumptions!C32"
    ws['E7'] = "33 months"
    
    # Formatting
    for row in range(3, 6):
        ws[f'B{row}'].style = 'currency'
    
    ws['B7'].style = 'currency'
    
    style_sheet(ws, styles, {'A': 25, 'B': 18, 'C': 15, 'D': 15, 'E': 15, 'F': 50})


# --- Main execution ---

def main():

    wb = setup_workbook()

    styles = apply_styles(wb)

    

    populate_readme(wb['README'], styles)

    populate_assumptions(wb['Assumptions'], styles)

    populate_summary(wb['Summary'], styles)

    populate_tranche_detail(wb['Tranche_Detail'], styles)

    populate_funnel_model(wb['Funnel_Model'], styles)

    populate_unit_economics(wb['Unit_Economics'], styles)

    populate_cash_runway(wb['Cash_Runway'], styles)

    populate_pay_triggers(wb['Pay_Triggers'], styles)

    populate_milestones(wb['Milestones'], styles)

    populate_scenarios(wb['Scenarios'], styles)

    populate_scenarios_output(wb['Scenarios_Output'], styles)

    populate_sales_comp(wb['Sales_Comp'], styles)

    populate_infra_estimates(wb['Infra_Estimates'], styles)

    populate_financial_summary(wb['Financial_Summary'], styles)

    populate_kpi_summary(wb['KPI_Summary'], styles)

    populate_funding_overview(wb['Funding_Overview'], styles)


    # Set 'Assumptions' as the active sheet

    wb.active = wb['Assumptions']

    

    # Save the workbook

    file_name = "BLU Use-of-Funds Model.xlsx"

    wb.save(file_name)

    print(f"Successfully created '{file_name}'")



if __name__ == "__main__":

    main()

